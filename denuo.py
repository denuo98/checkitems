#coding:utf-8
from Tkinter import *
import Tkinter
import xlrd
from openpyxl import load_workbook
import openpyxl
import time
import sys  
  
reload(sys)  
sys.setdefaultencoding('utf8')

root = Tk()
root.title('考研打卡')
#root.geometry('420x320')  # 设置窗口大小
root.geometry()
root.resizable(width=False, height=True)#禁止拉伸
#完成情况记录表，默认未完成
#flags=[False,False,False]
flags=[]
done_list = []
content_list =[]
count=1

def get_item_list_from_excel(item_file):
    item_list = []
    flags=[]
    excelFile = xlrd.open_workbook(item_file)
    sheet = excelFile.sheet_by_name('Sheet1')
    print(sheet.name, sheet.nrows, sheet.ncols)
    row=0
    while row<sheet.nrows :
        print type(sheet.cell(row, 0).value)
        if type(sheet.cell(row, 0).value)==float:#float类型的对象没有encode属性，所以表中纯数字要拎出来处理
            item_name  = str(sheet.cell(row, 0).value)
        else:
            item_name  = sheet.cell(row, 0).value.encode('utf-8')
            if sheet.cell(row, 1).value==1:
                item_flag  = True
            else:
                item_flag  = False
        '''添加表的其他属性
        item_start_time= sheet.cell(0, 1).value.encode('utf-8')
        '''
        #item_list.append((job_num, job_name))
        if item_flag==True:
            done_list.append(item_name)
        item_list.append(item_name)
        flags.append(item_flag)
        row+=1
    print item_list
    return item_list,flags

def write_item_list_to_excel(item_file):
    wb = load_workbook(item_file)
    wb_sheet = wb.get_sheet_by_name('Sheet1')
    clear=100
    index=0
    while index<clear:
        index+=1
        wb_sheet.cell(index+1, column = 1, value ='')
        #print str(done_list)+'2222222222222222222222'+str(content_list)
        wb_sheet.cell(index+1, column = 2, value ='')

    row=0#这个包里对excel表的行数从1开始
    while row<len(content_list):
        wb_sheet.cell(row+1, column = 1, value = content_list[row])
        #print str(done_list)+'11111111111111111111111'+str(content_list)
        wb_sheet.cell(row+1, column = 2, value = flags[row])

        row+=1
    wb.save(item_file)#保存
    return 0 

def exit_with_write(root):
    write_item_list_to_excel('items.xlsx')
    show_state()
    root.quit()


def click_done(item):
    loc=content_list.index(item)
    flags[loc]=not flags[loc]
    if flags[loc]:
        done_list.append(content_list[loc])
    else:
        done_list.remove(content_list[loc])
    '''
    str__='\n'
    c=1
    for item in done_list:
        print item
        str__+=str(c)+'_'+str(item)+'\n'
        c+=1
    lab_done['text']=str__
    '''
    check()
    show_state()
    return 0
def check():
    str__='\n'
    c=1
    for item in done_list:
            str__+=str(c)+'_'+str(item)+'\n'
            c+=1
    lab_done['text']=str__
    show_state()
    return 0

def click_del(item):
    loc=content_list.index(item)
    if flags[loc]:#若已经完成
        done_list.remove(item)#从完成列表删除注册信息
    del flags[loc]#从状态表删除注册信息
    content_list.remove(item)#从事项列表删除注册信息
    ck[loc].grid_forget()#删除对应的复选框和按钮
    btns_del[loc].grid_forget()
    btns_up[loc].grid_forget()
    #删除各组件在列表中的注册信息
    del ck[loc]
    del btns_del[loc]
    del btns_up[loc]
    show_state()
    return 0


def click_up(item):
    loc=content_list.index(item)
    if loc==0:#如果是第一个复选框，不能再上移，啥也不做
        show_state()
        return 0
    #在content_list和flags中上移
    content_list[loc],content_list[loc-1]=content_list[loc-1],content_list[loc]
    flags[loc],flags[loc-1]=flags[loc-1],flags[loc]
    global ck,btns_up,btns_del
    show_state('交换')
    #删除事项的复选框、删除键、上移键
    for item in ck:
        item.grid_forget()
    for item in btns_del:
        item.grid_forget()
    for item in btns_up:
        item.grid_forget()
    #清空注册表
    ck=[]
    btns_del=[]
    btns_up=[]
    #依据两个list更新视图
    update_ui()
    return 0
def update_ui():
    show_state()
    #依据新的事项列表新建复选框、按钮
    for item in content_list:
        ck.append(btn_build(item))
        btn_del,btn_up=btn_del_up_build(item)
        btns_del.append(btn_del)
        btns_up.append(btn_up)
        ck[content_list.index(item)].grid(row=content_list.index(item),sticky=W)
        btn_del.grid(row=content_list.index(item),column=2)
        btn_up.grid(row=content_list.index(item),column=3)
    return 0

def click_add():
    #添加输入框和确认键
    #调用添加事项函数
    global content_input,btn_confirm
    root_add=Tkinter.Tk()
    root_add.title('新建事项')
    root_add.geometry('250x55')
    #root_add.resizable(width=False, height=False)#禁止拉伸
    content_input=Entry(root_add,text='')
    btn_confirm=Button(root_add,text='添加',command=lambda:click_confirm(root_add))
    lab=Label(root_add,text='输入新事项：')
    lab.pack(anchor=W,side=LEFT)
    content_input.pack(side=LEFT)
    btn_confirm.pack(side=RIGHT)
    Label(root_add,text='~长度最好不要超过12个中文字符(●\'◡\'●)').pack(side=BOTTOM,anchor=W,before=lab)
    return 0
def click_confirm(root_add,item='xxx'):
    if item =='xxx':
        item=content_input.get()#从输入框获取事项
    print type(item)
    content_list.append(item)#添加到事项列表
    flags.append(False)#添加事项状态列表
    btn=btn_build(item)#创建按钮
    btn_del,btn_up=btn_del_up_build(item)
    ck.append(btn)#添加到按钮列表
    btns_del.append(btn_del)
    btns_up.append(btn_up)
    show_state('按钮注册')
    btn.grid(row=content_list.index(item),sticky=W)
    btn_del.grid(row=content_list.index(item),column=1)
    btn_up.grid(row=content_list.index(item),column=2)
    btn_confirm.grid_forget()
    content_input.grid_forget()
    root_add.destroy()
    return 0
#创建按钮对象，为什么用函数？因为在循环中，commander中函数参数总是会指定最后一个item，所以用函数封装一下
def btn_build(item):
    btn=Checkbutton(frm, text=item, command=lambda:click_done(item))
    if flags[content_list.index(item)]:
        btn.select()
    return btn
def btn_del_up_build(item):
    btn_del=Button(frm,text='del',fg='white',image=del_img,command=lambda:click_del(item))
    btn_up=Button(frm,text='up',fg='white',image=up_img,command=lambda:click_up(item))
    return btn_del,btn_up
def get_countdown():
    from datetime import datetime
#构造一个将来的时间
    future = datetime.strptime('2019-12-24 8:30:00','%Y-%m-%d %H:%M:%S')
#当前时间
    now = datetime.now()
#求时间差
    delta = future - now
    hour = delta.seconds/60/60
    minute = (delta.seconds - hour*60*60)/60
    seconds = delta.seconds - hour*60*60 - minute*60
    return str(delta.days)+'天'+ str(hour)+'小时'+str(minute)+'分钟'+ str(seconds)+'秒'
def show_state(str_=''):
    global count
    '''print "第"+str(str(count))+"次操作————》》》》》》"+str_
    print "##########逻辑事项注册信息###################"
    print content_list 
    print done_list
    print flags
    print '----------UI组件注册信息--------------------'
    print ck
    print btns_del
    print btns_up
    print '----------------事项完成情况-----------------'
    print type(lab_msg)'''
    count=count+1
def click_export():
    root_exp=Tkinter.Tk()
    root_exp.title('确认导出')
    root_exp.geometry('300x100')
    Label(root_exp,text='建议在当日所有事项都已列举完毕后再导出存档\n您确定要导出吗？\n(1小时内重复导出会覆盖上次导出结果)').pack()
    Button(root_exp,text='确认导出',command=lambda:exp_confirm(root_exp)).pack()
    return 0
def exp_confirm(root_exp):
    with open(time.strftime("%b_%d_%a_%H_%Y",time.localtime())+".txt","w") as f:
        f.write("存档时间 ："+time.ctime()+'\n')
        f.write("事项列表 ：\n")
        row=0
        while row<len(content_list):
            if flags[row]:
                f.write("完成_√\t\t\t")
            else:
                f.write("未成_×\t\t\t")
            f.write(content_list[row]+'\n')
            row+=1

    root_exp.destroy()


if __name__ == '__main__':

    #窗体控件
    # 标题显示
    lab = Label(root, text='今日事项：')
    lab.grid(row=0, column=0,sticky=W)

    lab_msg = Label(root, text="fininshed:")
    lab_msg.grid(row=0, column=1,sticky=W)

    frm1=Frame(root,borderwidth=1)
    frm1.grid(row=1,column=1)

    lab_done=Label(frm1,text='')
    lab_done.grid(row=0,column=0,sticky=NW)

    # 多选框
    frm = Frame(root,borderwidth=1)
    frm.grid(row=1,column=0)
    #处理图片
    del_img = PhotoImage(file = 'del.gif')
    up_img = PhotoImage(file = 'up.gif')
    #打卡键
    ck=[]
    btns_del=[]
    btns_up=[]
    content_list,flags=get_item_list_from_excel('items.xlsx')
    for item in content_list:
        ck.append(btn_build(item))
        btn_del,btn_up=btn_del_up_build(item)
        btns_del.append(btn_del)
        btns_up.append(btn_up)
        ck[content_list.index(item)].grid(row=content_list.index(item),sticky=NW)
        btn_del.grid(row=content_list.index(item),column=1,sticky=NW)
        btn_up.grid(row=content_list.index(item),column=2,sticky=NW)
        show_state()
    check()
    btn_add=Button(root,text='新建',command=click_add)
    btn_add.grid(row=2,column=0,sticky=E)
    btn_export=Button(root,text='导出',command=click_export)
    btn_export.grid(row=3,column=1,sticky=E)
    btn_exit=Button(root,text='保存/退出',command=lambda:exit_with_write(root))
    btn_exit.grid(row=4,column=1,sticky=E)
    lab_time=Label(root,text='NOW : '+time.strftime("%b %d %a %Y",time.localtime()))
    lab_time.grid(row=5,column=0)
    lab_contact=Label(root,text='Contact Me : yehy2016@lzu.edu.cn')
    lab_contact.grid(row=5,column=1,sticky=E)
    Label(root,text='考研路漫漫，与君相伴2019~').grid(row=6,column=1)
    lab_countdown=Label(root,text='TIME LEFT : '+get_countdown())
    lab_countdown.grid(row=6,column=0)
    root.mainloop()

#事项长度固定                     完成 无意义 因为换行在label内，固定长度以后，ui显示还是对不齐 函数体已经被注释掉
#list写入excel_xlsx文件保存       完成
#加入时间线                       完成 因为不是数据库管理，添加时间线有些困难，所以改成了“导出功能”（觉得导入不必要，过去的就过去了）
#读取和写入flags，在excel         完成  done_list不需要保存，done_list在逻辑上可以被content_list和flags确定，功能上只是作为字符串显示而已
#加入倒计时                       完成
#加入邮件|公众号提醒
#开机自启
#后台运行



'''写入excel的面向对象方法：
# class Write_excel(object):
#     修改excel数据
#     def __init__(self, filename):
#         self.filename = filename
#         self.wb = load_workbook(self.filename)
#         self.ws = self.wb.active  # 激活sheet
#
#     def write(self, row_n, col_n, value):
#         写入数据，如(2,3,"hello"),第二行第三列写入数据hello
#         self.ws.cell(row_n, col_n,value )
#         self.wb.save(self.filename)
#
# we = Write_excel("mylogintest.xlsx")
# we.write(2,2,'pass3')
'''
